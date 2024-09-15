

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import render
from django.views import View
from test_1.models import *
from io import BytesIO
import openpyxl



# Задание №2 часть вторая
class Zadanie_2_2_View(LoginRequiredMixin, View):
    model = User
    template_name = "test_1/zadanie_1/zadanie_1.html"

    def get(self, request, *args, **kwargs):
        # Получим пользователя
        polzovatel = self.request.user

        # В данном случае не было задачи описывать как именно игрок проходит уровень. Просто создадим переменную
        # которая будет означать, что игрок прошел новый уровень и ему положен приз
        prize_add = 1
        if prize_add:
            book = Metods_2.CSV_1(polzovatel)
            print(book)
            file_stream = BytesIO()
            book.save(file_stream)

            response = HttpResponse(content=file_stream.getvalue(), content_type='application/x-download')
            response.content_type = "application/octet-stream;"
            response["Content-Disposition"] = f"attachment; filename=table.csv"
            return response
        massiv_10 = PlayerLevel.objects.all().prefetch_related('level__level_prize_1')

        slovar_itog = {'massiv_10': massiv_10,}

        return render(request, self.template_name, slovar_itog)


class Metods_2:

    @staticmethod
    def CSV_1(polzovatel):

        # Cоздадим книгу эксель
        book = openpyxl.Workbook()
        book.remove(book.active)
        book_active = book.create_sheet('Выгрузка 1')
        # Заполним заголовки
        Metods_2.Zagolovki(book_active)
        # Получим массив с данными
        spisok_1 = Metods_2.Massiv(book_active)
        dlina_1 = len(spisok_1) + 1
        # Внесем данные в таблицу
        for idx, row in enumerate(book_active.iter_rows(min_row=2, max_col=4, max_row=dlina_1)):
            for idex, cell in enumerate(row):
                cell.value = spisok_1[idx][idex]

        book.save('Vigruzka_1.csv')
        book.close()

        return book


    # Заголовки
    @staticmethod
    def Zagolovki(a):
        a['A1'] = 'id Игрока'
        a['B1'] = 'Название уровня'
        a['C1'] = 'Пройден ли уровень'
        a['D1'] = 'Полученный приз за уровень'


    # Массив с данными.
    @staticmethod
    def Massiv(a):
        # Сделаем словарь со значениями призов и уровней
        slovar_prizov = {}
        for i in LevelPrize.objects.select_related('prize'):
            slovar_prizov[i.pk] = i.prize.title if i.prize else ''

        massiv = PlayerLevel.objects.select_related('player', 'level')
        spisok_1 = []
        for idx, i in enumerate(PlayerLevel.objects.select_related('player', 'level')):
            idx_2 = idx + 1

            player_pk = i.player.pk if i.player else ''
            level_title = i.level.title if i.level else ''
            priz = slovar_prizov[i.level.pk] if i.level.pk in slovar_prizov else ''

            spisok_1_2 = [player_pk, level_title, i.is_completed, priz]
            spisok_1.append(spisok_1_2)

        return spisok_1














